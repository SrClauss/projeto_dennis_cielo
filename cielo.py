from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from time import sleep
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import requests
import undetected_chromedriver as uc



with open("config.json", "r") as file:
    config = json.load(file)
LOGIN = config["login"]
SENHA = config["senha"]
URL = config["url"]
estabelecimentos_vazios = []
errors = []
wait_input = config["wait_input"]


estabelecimento = "1012248698"
def recarregue(driver):
    try:
        h1s = WebDriverWait(driver, 5).until(EC.presence_of_all_elements_located((By.TAG_NAME, "h1")))
        return h1s
    except Exception as e:
        driver.get("https://minhaconta2.cielo.com.br/site/vendas/resumo/cielo")
        return recarregue(driver)


def extract_filial(i, estabelecimento, driver):
    pesquisar_estabelecimentos = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@placeholder="Pesquisar estabelecimentos"]')))
    pesquisar_estabelecimentos.click()
    ActionChains(driver).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
    ActionChains(driver).send_keys(Keys.BACKSPACE).perform()
    ActionChains(driver).send_keys(estabelecimento).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    list_item = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "list-item")))
    cnpj = list_item[i].text.split("\n")[-1]
    list_item[i].click()
    button_acessar = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[text()="Acessar"]')))
    button_acessar.click()
    sleep(0.5)
    driver.get("https://minhaconta2.cielo.com.br/site/vendas/resumo/cielo")

    h1s = recarregue(driver)


    
    valor_ontem = h1s[3]
    valor_hoje = h1s[5]
    while valor_ontem.text == "" or valor_hoje.text == "":
        h1s = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.TAG_NAME, "h1")))
        valor_ontem = h1s[3]
        valor_hoje = h1s[5]
    if h1s[3].text == "Hoje":
        print("Erro ao carregar página: ", estabelecimento)
        global errors
        errors.append(estabelecimento)
        return None

    if valor_ontem.text == "-":
        valor_ontem = 0
    else:
        valor_ontem = float(valor_ontem.text.replace("R$ ", "").replace(".", "").replace(",", "."))
    if valor_hoje.text == "-":
        valor_hoje = 0
    else:
        valor_hoje = float(valor_hoje.text.replace("R$ ", "").replace(".", "").replace(",", "."))
    driver.execute_script("window.scrollTo(0, 300);")
    calendar = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//div[@class="common-icon"]//following-sibling::input')))
    calendar.click()
    button_historico = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[text()="Histórico"]')))
    button_historico.click()
    button_ultimos_30_dias = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//span[text()="Últimos 30 dias"]//parent::button')))
    button_ultimos_30_dias.click()
    button_confirmar = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[text()="Confirmar"]')))
    button_confirmar.click()
    button_visualizar_totais = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[text()="Visualizar totais"]')))
    button_visualizar_totais.click()
    floating_header = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "floating-header")))
    WebDriverWait(driver, 10).until(lambda driver: floating_header.text != "")
    bruto_30, taxa_30, liquido_30 = [float(valor.split('\n')[0].replace('.', '').replace(',', '.')) for valor in floating_header.text.split("R$ ")[1:]]
    result = {
        "estabelecimento": estabelecimento,
        "cnpj": cnpj,
        "valor_ontem": valor_ontem,
        "valor_hoje": valor_hoje,
        "bruto_30": bruto_30,
        "taxa_30": taxa_30,
        "liquido_30": liquido_30
    }

    print(result)
    return result

def login():

    driver = webdriver.Chrome()
    driver.get(URL)
    driver.maximize_window()
    bt_other_access = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "bt-other-access")))
    bt_other_access.click()
    ActionChains(driver).send_keys(Keys.TAB).perform()
    ActionChains(driver).send_keys(LOGIN).perform()
    ActionChains(driver).send_keys(Keys.TAB).perform()
    ActionChains(driver).send_keys(SENHA).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    if wait_input:
        input("Pressione Entar para continuar...")
      
    
    return driver
def login_undetectable():
    driver = uc.Chrome()
    driver.get(URL)
    driver.maximize_window()
    bt_other_access = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "bt-other-access")))
    bt_other_access.click()
    ActionChains(driver).send_keys(Keys.TAB).perform()
    ActionChains(driver).send_keys(LOGIN).perform()
    ActionChains(driver).send_keys(Keys.TAB).perform()
    ActionChains(driver).send_keys(SENHA).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    return driver
    
def extract_estabelecimento(estabelecimento, driver):
 
   
    pesquisar_estabelecimentos = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@placeholder="Pesquisar estabelecimentos"]')))    
    pesquisar_estabelecimentos.click()
    ActionChains(driver).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
    ActionChains(driver).send_keys(Keys.BACKSPACE).perform()
    ActionChains(driver).send_keys(estabelecimento).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    list_item_len = 0
   
    try:
        list_item = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "list-item")))
        list_item_len =  len(list_item)
    except Exception as e:
        global estabelecimentos_vazios
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "empty-search")))
            estabelecimentos_vazios.append(estabelecimento)
            print("erro de estabelencimento vazio")
            return None
        except Exception as e:
            print(f"erro geral: {estabelecimento}")
            global errors
            errors.append(estabelecimento)
            return None

    result = []


    
    for i in range(list_item_len):
        result.append(extract_filial(i, estabelecimento, driver))
        trocar_estabelecimento = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//span[text()="Trocar estabelecimento"]')))
        trocar_estabelecimento.click()

    return result
    



# Supondo que a função login() e extract_estabelecimento() já estejam definidas

def dividir_array(arr, n):
    """Divide um array em n partes, todas iguais exceto a última que recebe os restantes."""

    tamanho_base = len(arr) // n  # Tamanho das partes, exceto a última
    resultado = []

    for i in range(n):
        inicio = i * tamanho_base
        fim = inicio + tamanho_base
        if i == n - 1:  # Última parte, pega até o final
            fim = len(arr)
        resultado.append(arr[inicio:fim])
    
    return resultado
def multi_estabelecimentos(estabelecimentos, driver):
    result = []
    for i in range(len(estabelecimentos)):
        result.append(extract_estabelecimento(estabelecimentos[i], driver))
    return result
if __name__ == "__main__":
    data = []

    hora_atual = datetime.fromisoformat(datetime.now().isoformat())

   
    if config["undetected_mode"] == False:
        hora = hora_atual.hour
        workers = 3
        if hora >= 0 and hora < 8:
            workers = config["workers_0h_8h"]
        elif hora >= 8 and hora < 10:
            workers = config["workers_8h_10h"]
        elif hora >= 10 and hora < 12:
            workers = config["workers_10h_12h"]
        elif hora >= 12 and hora < 14:
            workers = config["workers_12h_14h"]
        elif hora >= 14 and hora < 16:
            workers = config["workers_14h_16h"]
        elif hora >= 16 and hora < 18:
            workers = config["workers_16h_18h"]
        elif hora >= 18 and hora < 20:
            workers = config["workers_18h_20h"]
        elif hora >= 20 and hora < 22:
            workers = config["workers_20h_22h"]
        else:
            workers = config["workers_22h_0h"]
        
    



        workbook_base = openpyxl.load_workbook("Base.xlsx")
        sheet_base = workbook_base.active
        estabelecimentos = [cell.value for cell in sheet_base["A"] if cell.value][1:]
        estabelecimentos_divididos = dividir_array(estabelecimentos, workers)
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [executor.submit(login) for i in range(workers)]
            drivers = [future.result() for future in as_completed(futures)]
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [executor.submit(multi_estabelecimentos, estabelecimentos_divididos[i], drivers[i]) for i in range(workers)]
            for future in as_completed(futures):
                data.extend(future.result())
                
        for driver in drivers:
            driver.quit()

    else:
        workbook_base = openpyxl.load_workbook("Base.xlsx")
        sheet_base = workbook_base.active
        estabelecimentos = [cell.value for cell in sheet_base["A"] if cell.value][1:]
        driver = login_undetectable()
        data = multi_estabelecimentos(estabelecimentos, driver)
        driver.quit()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Estabelecimento", "CNPJ", "Valor Ontem", "Valor Hoje", "Bruto 30", "Taxa 30", "Liquido 30"])  
    data = [d for d in data if d != None]
    soma = 0
    for d in data:
        for t in d:
            print("inserido:", soma)
            soma += 1
            sheet.append([t["estabelecimento"], t["cnpj"], t["valor_ontem"], t["valor_hoje"], t["bruto_30"], t["taxa_30"], t["liquido_30"]])

    
    sheet_errors = workbook.create_sheet("Erros")
    sheet_errors.append(["Estabelecimento"])
    for error in errors:
        sheet_errors.append([error])
    sheet_vazios = workbook.create_sheet("Estabelecimentos Vazios")
    sheet_vazios.append(["Estabelecimento"])
    for vazio in estabelecimentos_vazios:
        sheet_vazios.append([vazio])
    file_name = f"Relatório {datetime.now().strftime('%d-%m-%Y %H-%M-%S')}.xlsx"
    workbook.save(file_name)


    print("Relatório gerado com sucesso")
    




    




        